from openpyxl.styles import PatternFill
from openpyxl import load_workbook

# Save DataFrame to Excel
combined_df.to_excel(writer, index=False, sheet_name=sheet_name)
writer.save()

# Load the Excel file
wb = load_workbook(writer.path)
ws = wb[sheet_name]

# Red fill style
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Header row (usually first row)
header = [cell.value for cell in ws[1]]

# Define which columns' headers you want to color
target_columns = ['Information', 'Underwriter (calculation that already has been populated on BP filled underwriter)', 'Comment (UW)']

# Loop through and color only header cells
for col_idx, col_name in enumerate(header, start=1):  # Excel is 1-indexed
    if col_name in target_columns:
        ws.cell(row=1, column=col_idx).fill = red_fill

# Save changes
wb.save(writer.path)
