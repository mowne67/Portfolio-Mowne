from openpyxl import load_workbook

def reorder_rows_in_app_sheets(file_path: str, required_fields: list[str]) -> None:
    wb = load_workbook(file_path)

    for ws in wb.worksheets:
        if not ws.title.startswith("App"):
            continue

        # Step 1: Transpose sheet (rows → columns)
        transposed = []
        for col in range(1, ws.max_column + 1):
            transposed.append([
                ws.cell(row=row, column=col).value for row in range(1, ws.max_row + 1)
            ])

        # Step 2: Build map from field name → column (which was a row originally)
        field_to_column = {
            col[0]: col for col in transposed[1:]  # skip column A
        }

        # Step 3: Rebuild transposed matrix with reordered fields
        new_transposed = [transposed[0]]  # keep header column (field labels)
        for field in required_fields:
            if field in field_to_column:
                new_transposed.append(field_to_column[field])
            else:
                # If field not found, append blank row with just the label
                new_transposed.append([field] + [""] * (len(transposed[0]) - 1))

        # Step 4: Transpose back (columns → rows)
        final_data = list(zip(*new_transposed))

        # Step 5: Clear sheet and write back
        ws.delete_rows(1, ws.max_row)
        for i, row in enumerate(final_data, start=1):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j).value = value

    # Step 6: Save changes to the workbook
    wb.save(file_path)
